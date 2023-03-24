# -*- coding: utf-8 -*-
"""
Created on Fri Mar 10 11:04:13 2023

@author: siddh
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Mar 13 22:36:34 2023

@author: siddh
"""
import instaloader
import openpyxl


L = instaloader.Instaloader()


hashtag = input("Enter the hashtag to search for: ")


hashtag_obj = instaloader.Hashtag.from_name(L.context, hashtag)


try:
    wb = openpyxl.load_workbook("instagram_data_new.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Username", "Userid", "Total posts", "Followers", "Following", "Bio", "Profile link"])


max_posts = 30


for post in hashtag_obj.get_posts():
    try:
        username = post.owner_username
        profile = instaloader.Profile.from_username(L.context, username)
        
        
        if profile.followers >= 2000:
            
            ws.append([profile.username, profile.userid, profile.mediacount, profile.followers, profile.followees, profile.biography, profile.external_url])
            
            
            print(f"Added {profile.username} to the Excel file.")

            
            if ws.max_row >= max_posts:
                break
    except KeyError:
        print(f"Skipping post {post} due to missing data")
        continue

wb.save("instagram_data_new.xlsx")
